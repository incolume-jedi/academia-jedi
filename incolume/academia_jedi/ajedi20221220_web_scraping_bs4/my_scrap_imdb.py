import logging
from pathlib import Path
from bs4 import BeautifulSoup
import requests
import openpyxl

logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s;%(levelname)-8s;%(name)s;'
    '%(module)s;%(funcName)s;%(message)s',
)


def scraping_ranking(
    url: str = '', excel_output: (str, Path) = '', columns_name: list = None
) -> bool:
    excel_output = Path(excel_output or 'my_IMDB_Movies_Ratings.xlsx')
    excel = openpyxl.Workbook()
    logging.debug(excel.sheetnames)
    sheet = excel.active
    sheet.title = 'Top rate movies'
    logging.debug(excel.sheetnames)
    sheet.append(
        columns_name
        or [
            'Movie Rank',
            'Movie name',
            'Year of release',
            'IMDB Ranking',
            'Poster',
        ]
    )

    url = url or 'https://www.imdb.com/chart/top'
    try:
        req = requests.get(url)
        req.raise_for_status()

        soup = BeautifulSoup(req.content, 'html.parser')
        # logging.debug(soup)
        movies = soup.find('tbody', class_='lister-list').find_all('tr')
        logging.debug('%s; %s', len(movies), movies)
        for movie in movies:
            # print(movie)
            name = movie.find('td', class_='titleColumn').a.text
            rank = (
                movie.find('td', class_='titleColumn')
                .get_text(strip=True)
                .split('.')[0]
            )
            year = movie.find('td', class_='titleColumn').span.text.strip('()')
            rating = movie.find(
                'td', class_='ratingColumn imdbRating'
            ).strong.text
            poster = movie.find('td', class_='posterColumn').img['src']
            logging.debug(
                '{ rank: %s, name: %s,year: %s, rating: %s, poster: %s}',
                rank,
                name,
                year,
                rating,
                poster,
            )
            sheet.append([rank, name, year, rating, poster])
            # break
    except requests.exceptions.HTTPError as e:
        logging.error(e)

    excel.save(excel_output.as_posix())
    return excel_output.is_file()
