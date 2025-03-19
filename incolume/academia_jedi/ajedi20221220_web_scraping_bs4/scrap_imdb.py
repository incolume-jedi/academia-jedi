"""Baseado em [Web Scraping in Python using Beautiful Soup.

Writing a Python program to Scrape IMDB website]
(https://www.youtube.com/watch?v=LCVSmkyB4v8).
"""

import logging

import openpyxl
import requests
from bs4 import BeautifulSoup
from config import settings

excel = openpyxl.Workbook()
logging.debug(excel.sheetnames)
sheet = excel.active
sheet.title = 'Top rate movies'
logging.debug(excel.sheetnames)
sheet.append(
    ['Movie Rank', 'Movie name', 'Year of release', 'IMDB Ranking', 'Poster'],
)

url = 'https://www.imdb.com/chart/top'
try:
    req = requests.get(url, timeout=settings.timeout)
    req.raise_for_status()

    soup = BeautifulSoup(req.content, 'html.parser')
    movies = soup.find('tbody', class_='lister-list').find_all('tr')
    logging.debug('%s; %s', len(movies), movies)
    for movie in movies:
        name = movie.find('td', class_='titleColumn').a.text
        rank = (
            movie.find('td', class_='titleColumn')
            .get_text(strip=True)
            .split('.')[0]
        )
        year = movie.find('td', class_='titleColumn').span.text.strip('()')
        rating = movie.find('td', class_='ratingColumn imdbRating').strong.text
        poster = movie.find('td', class_='posterColumn').img['src']
        logging.debug(
            '{ rank: %s, name: %s,year: %s, rating: %s, poster: %s}',
            rank,
            name,
            year,
            rating,
            poster,
        )
        sheet.append([rank, name, year, rating, poster])
except requests.exceptions.HTTPError:
    msg = 'Erro HTTP.'
    logging.exception(msg)

excel.save('IMDB_Movies_Ratings.xlsx')
