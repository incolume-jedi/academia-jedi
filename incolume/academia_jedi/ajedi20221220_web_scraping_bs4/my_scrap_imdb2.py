"""Module scrap imdb2."""

from __future__ import annotations

import sys
from dataclasses import dataclass
from pathlib import Path

import openpyxl
import requests
from bs4 import BeautifulSoup
from config import settings
from incolume.academia_jedi import logger

if sys.version_info < (3, 11):
    from typing_extensions import Self
else:
    from typing import Self


@dataclass
class Movie:
    """Dataclass movie."""

    rank: int
    name: str
    year: int
    rating: str
    poster: str


@dataclass
class ScrapingIMDB:
    """Class ScrapingIMDB."""

    def __init__(
        self,
        url: str = '',
        excel_output: (str | Path) = '',
        sheet_title: str = '',
        columns_name: list | None = None,
    ) -> None:
        """Init class."""
        self.url = url or 'https://www.imdb.com/chart/top'
        self.excel_output = Path(excel_output or 'my_IMDB_Movies_Ratings.xlsx')
        self.sheet_title = sheet_title or 'Top rate movies'
        self.columns_name = columns_name or [
            'Movie Rank',
            'Movie name',
            'Year of release',
            'IMDB Ranking',
            'Poster',
        ]
        self.req = None
        self.soup = None
        self.movies = []

    def connect(self, url: str = '') -> Self:
        """Connect."""
        url = url or self.url
        try:
            self.req = requests.get(url, timeout=settings.timeout)
            self.req.raise_for_status()
        except requests.exceptions.HTTPError:
            msg = 'Erro HTTP.'
            logger.exception(msg)
        return self

    def get_soup(self):
        """Get soup."""
        self.soup = BeautifulSoup(self.req.content, 'html.parser')
        return self

    def get_movies(self):
        """Get movie."""
        try:
            movies = self.soup.find('tbody', class_='lister-list').find_all(
                'tr',
            )
        except AttributeError:
            return self

        for movie in movies:
            obj = Movie(
                rank=movie.find('td', class_='titleColumn')
                .get_text(strip=True)
                .split('.')[0],
                name=movie.find('td', class_='titleColumn').a.text,
                year=movie.find('td', class_='titleColumn').span.text.strip(
                    '()',
                ),
                rating=movie.find(
                    'td',
                    class_='ratingColumn imdbRating',
                ).strong.text,
                poster=movie.find('td', class_='posterColumn').img['src'],
            )
            logger.debug(
                '{ rank: %s, name: %s,year: %s, rating: %s, poster: %s}',
                *obj.__dict__.values(),
            )
            self.movies.append(obj)
        logger.debug('Quantidade encontrada: %s', len(self.movies))
        return self

    def save_excel(
        self,
        *,
        excel_output: (str | Path) = '',
        columns_name: list | None = None,
        **kwargs: dict,
    ) -> bool:
        """Save excel."""
        try:
            if columns_name is None:
                columns_name = []
            excel_output = Path(excel_output or self.excel_output)
            excel = openpyxl.Workbook()
            logger.debug(excel.sheetnames)

            sheet = excel.active
            sheet.title = kwargs.get('sheet_title') or self.sheet_title
            logger.debug(excel.sheetnames)

            sheet.append(columns_name or self.columns_name)
            [
                sheet.append(list(movie.__dict__.values()))
                for movie in self.movies
            ]
            excel.save(excel_output.as_posix())
        except AttributeError:
            return False
        return True

    def scraping(self, **kwargs):
        """Scraping."""
        url = kwargs.get('url') or self.url
        excel_output = Path(kwargs.get('excel_output') or self.excel_output)
        sheet_title = kwargs.get('sheet_title') or self.sheet_title
        columns_name = kwargs.get('columns_name') or self.columns_name
        return (
            self.connect(url)
            .get_soup()
            .get_movies()
            .save_excel(
                excel_output=excel_output,
                columns_name=columns_name,
                sheet_title=sheet_title,
            )
        )
